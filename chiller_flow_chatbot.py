import streamlit as st
import math
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

# === OneDrive Excel File Path ===
EXCEL_PATH = r"C:\Users\tsvaevq\OneDrive - Volkswagen AG\SmartChillerChatbotExcel.xlsx"

# === Streamlit Config ===
st.set_page_config(page_title="Air Flow Calculator", page_icon="💨", layout="centered")
st.title("🤖 Smart Chiller Air Flow Assistant")
st.write("👋 Hello! How can I help you today?")

# === Session State Initialization ===
default_keys = {
    'step': "start", 'equipment': None, 'shape': None, 'unit': None,
    'dimensions': {}, 'velocities': [], 'vel_unit': None, 'vel_count': 0, 'result': None
}
for k, v in default_keys.items():
    if k not in st.session_state:
        st.session_state[k] = v


# === Helper Functions ===
def convert_length(value, from_unit):
    conv = {"m": 1, "cm": 0.01, "foot": 0.3048, "inches": 0.0254}
    return value * conv[from_unit]

def convert_velocity(value, from_unit):
    conv = {
        "m/s": 1, "cm/s": 0.01, "ft/s": 0.3048, "inch/s": 0.0254,
        "m/min": 1/60, "cm/min": 0.01/60, "ft/min": 0.3048/60, "inches/min": 0.0254/60
    }
    return value * conv[from_unit]

def calc_area(shape, dims_m):
    if shape == "Square":
        return dims_m['length'] * dims_m['breadth']
    elif shape == "Round":
        return math.pi * (dims_m['diameter'] / 2) ** 2

def calc_flow(vels_m, area_m2):
    avg_vel = sum(vels_m) / len(vels_m)
    flow = avg_vel * area_m2
    return avg_vel, flow

def append_to_excel(data_dict):
    """Append the results to shared Excel file (create if missing)."""
    df_new = pd.DataFrame([data_dict])

    if os.path.exists(EXCEL_PATH):
        try:
            book = load_workbook(EXCEL_PATH)
            writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            start_row = writer.sheets['Sheet1'].max_row
            df_new.to_excel(writer, index=False, header=False, startrow=start_row, sheet_name='Sheet1')
            writer.close()
        except Exception as e:
            st.error(f"⚠️ Error updating Excel file: {e}")
    else:
        df_new.to_excel(EXCEL_PATH, index=False)
        st.info("📘 Excel file created successfully!")


# === Step Flow ===
if st.session_state.step == "start":
    if st.button("🔹 Calculate Air Flow"):
        st.session_state.step = "choose_equipment"
        st.rerun()

elif st.session_state.step == "choose_equipment":
    st.write("For which equipment would you like to calculate flow?")
    equipment = st.radio("Select Equipment", ["Cooling Tower", "Ventilation Unit"], horizontal=True)
    if st.button("Next ➡️"):
        st.session_state.equipment = equipment
        st.session_state.step = "choose_shape"
        st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "start"
        st.rerun()

elif st.session_state.step == "choose_shape":
    st.write("Specify the surface dimension of measured area")
    shape = st.radio("Select Shape", ["Square", "Round"], horizontal=True)
    if st.button("Next ➡️"):
        st.session_state.shape = shape
        st.session_state.step = "choose_unit"
        st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "choose_equipment"
        st.rerun()

elif st.session_state.step == "choose_unit":
    st.write("Select the measurement unit for surface area:")
    unit = st.radio("Measurement Unit", ["m", "cm", "foot", "inches"], horizontal=True)
    if st.button("Next ➡️"):
        st.session_state.unit = unit
        st.session_state.step = "enter_dimensions"
        st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "choose_shape"
        st.rerun()

elif st.session_state.step == "enter_dimensions":
    shape = st.session_state.shape
    unit = st.session_state.unit
    if shape == "Square":
        length = st.number_input(f"Enter Length ({unit})", min_value=0.0, step=0.1)
        breadth = st.number_input(f"Enter Breadth ({unit})", min_value=0.0, step=0.1)
        if st.button("Next ➡️"):
            st.session_state.dimensions = {"length": length, "breadth": breadth}
            st.session_state.step = "enter_velocity_count"
            st.rerun()
    else:
        diameter = st.number_input(f"Enter Diameter ({unit})", min_value=0.0, step=0.1)
        if st.button("Next ➡️"):
            st.session_state.dimensions = {"diameter": diameter}
            st.session_state.step = "enter_velocity_count"
            st.rerun()

    if st.button("⬅️ Go Back"):
        st.session_state.step = "choose_unit"
        st.rerun()

elif st.session_state.step == "enter_velocity_count":
    count = st.number_input("Enter how many velocity readings you will provide", min_value=1, step=1)
    if st.button("Next ➡️"):
        st.session_state.vel_count = count
        st.session_state.step = "choose_velocity_unit"
        st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "enter_dimensions"
        st.rerun()

elif st.session_state.step == "choose_velocity_unit":
    st.write("Select the velocity unit:")
    vel_unit = st.radio(
        "Velocity Unit",
        ["m/s", "cm/s", "ft/s", "inch/s", "m/min", "cm/min", "ft/min", "inches/min"],
        horizontal=False
    )
    if st.button("Next ➡️"):
        st.session_state.vel_unit = vel_unit
        st.session_state.velocities = []
        st.session_state.step = "enter_velocities"
        st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "enter_velocity_count"
        st.rerun()

elif st.session_state.step == "enter_velocities":
    vel_unit = st.session_state.vel_unit
    count = st.session_state.vel_count
    st.write(f"Enter {count} velocity readings in **{vel_unit}**:")
    next_idx = len(st.session_state.velocities) + 1
    if next_idx <= count:
        val = st.number_input(f"Velocity reading #{next_idx}", min_value=0.0, step=0.1)
        if st.button("Add Reading"):
            st.session_state.velocities.append(val)
            st.rerun()
    if len(st.session_state.velocities) == count:
        if st.button("Review ➡️"):
            st.session_state.step = "review_table"
            st.rerun()
    if st.button("⬅️ Go Back"):
        st.session_state.step = "choose_velocity_unit"
        st.rerun()

elif st.session_state.step == "review_table":
    st.write("Review your entered velocity readings:")
    df = pd.DataFrame(st.session_state.velocities, columns=["Velocity"])
    edited_df = st.data_editor(df, num_rows="dynamic", key="vel_edit")
    st.session_state.velocities = edited_df["Velocity"].tolist()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("➕ Add Extra Reading"):
            st.session_state.velocities.append(0.0)
            st.rerun()
    with col2:
        if st.button("🔙 Go Back"):
            st.session_state.step = "enter_velocities"
            st.rerun()
    with col4:
        if st.button("✅ Proceed"):
            st.session_state.step = "result"
            st.rerun()

elif st.session_state.step == "result":
    # --- Calculations ---
    dims_m = {k: convert_length(v, st.session_state.unit) for k, v in st.session_state.dimensions.items()}
    area_m2 = calc_area(st.session_state.shape, dims_m)
    vels_m = [convert_velocity(v, st.session_state.vel_unit) for v in st.session_state.velocities]
    avg_m, flow_m3s = calc_flow(vels_m, area_m2)

    flow_m3min = flow_m3s * 60
    flow_m3hr = flow_m3s * 3600

    # --- Display Results ---
    st.success(f"""
    ✅ **Calculation Complete!**
    - *Equipment:* {st.session_state.equipment}  
    - Shape: {st.session_state.shape}  
    - Surface Area: {area_m2:.3f} m²  
    - Average Velocity: {avg_m:.3f} m/s  
    - Air Flow Rate:  
        • {flow_m3s:.4f} m³/s  
        • {flow_m3min:.4f} m³/min  
        • {flow_m3hr:.4f} m³/hr
    """)

    # --- Save Results to Excel ---
    result_data = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Equipment": st.session_state.equipment,
        "Shape": st.session_state.shape,
        "Surface_Area_m2": area_m2,
        "Avg_Velocity_m/s": avg_m,
        "Flow_m3/s": flow_m3s,
        "Flow_m3/min": flow_m3min,
        "Flow_m3/hr": flow_m3hr,
        "Velocities": ", ".join(map(str, st.session_state.velocities))
    }
    append_to_excel(result_data)
    st.success("📊 Result saved to OneDrive Excel successfully!")

    # --- Restart ---
    if st.button("🔄 Start New Calculation"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()
